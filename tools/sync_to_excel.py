r"""Trägt die in der Stempeluhr-PWA erfassten Tage direkt in die
institutionelle Excel-Arbeitszeiterfassung ein.

Holt alle Tage eines Profils vom Pi (GET /api/{person}/days) und schreibt
Kommt/Geht-Zeiten, bezahlte Abwesenheit, Abwesenheitsgrund und Bemerkung in
die passende Zeile der Datei "Arbeitszeiterfassung <Jahr> ....xlsx"
(Tabellenblatt "Arbeitszeiterfassung", Zeile 4 = 1. Januar, danach ein Tag
pro Zeile). Formelzellen (Datum, Wochentag, Soll-/Ist-Arbeitszeit, Saldo)
werden nicht angerührt – nur die reinen Eingabespalten G–N, O, P, T.

Vor jedem Schreibvorgang wird automatisch eine Backup-Kopie der Datei mit
Zeitstempel angelegt (z.B. "...xlsx.backup-20260814-101500").

Nur Tage mit tatsächlichen Daten (Stempel, Absenz oder Bemerkung) werden
geschrieben – Tage ohne App-Eintrag bleiben in der Excel-Datei unangetastet,
falls dort z.B. schon manuell etwas eingetragen wurde.

Nutzung:
    python tools/sync_to_excel.py                    # schreibt direkt
    python tools/sync_to_excel.py --dry-run           # zeigt nur, was passieren würde
    python tools/sync_to_excel.py --person nadja --file "C:\...\Nadjas Zeiterfassung.xlsx"
"""

import argparse
import datetime as dt
import shutil
import sys
from pathlib import Path

import openpyxl
import requests

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass  # ältere Python-Version ohne reconfigure – Umlaute evtl. verstümmelt, sonst egal

DEFAULT_URL = "https://pi5.tail0fe4c7.ts.net"
DEFAULT_PERSON = "michael"
DEFAULT_FILE = (
    r"C:\Users\micha\Desktop\Privat\Ensis\Zeiterfassung"
    r"\Arbeitszeiterfassung 2026 ab 01.09.2026 - Michi Näf.xlsx"
)

SHEET = "Arbeitszeiterfassung"
FIRST_ROW = 4  # entspricht dem 1. Januar des Jahres in Berechnung!A3
COL_KOMMT_GEHT = {
    0: 7,   # G  Kommt 1
    1: 8,   # H  Geht 1
    2: 9,   # I  Kommt 2
    3: 10,  # J  Geht 2
    4: 11,  # K  Kommt 3
    5: 12,  # L  Geht 3
    6: 13,  # M  Kommt 4
    7: 14,  # N  Geht 4
}
COL_ABWESENHEIT_STD = 15    # O
COL_ABWESENHEIT_GRUND = 16  # P
COL_BEMERKUNG = 20          # T

# Muss exakt der Dropdown-Liste in P371:P379 der Excel-Vorlage entsprechen.
VALID_GRUENDE = {"Ferien", "krank", "Unfall", "geschäftlich", "Weiterbild'g"}


def fetch_days(base_url: str, person: str) -> dict:
    url = f"{base_url}/api/{person}/days"
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    return resp.json()


def parse_time(hhmm: str) -> dt.time:
    h, m = hhmm.split(":")
    return dt.time(int(h), int(m))


def has_content(day: dict) -> bool:
    return bool(day.get("punches")) or bool(day.get("abwesenheitStd")) or bool((day.get("bemerkung") or "").strip())


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--person", default=DEFAULT_PERSON, help=f"Profilname auf dem Server (Standard: {DEFAULT_PERSON})")
    ap.add_argument("--file", default=DEFAULT_FILE, help="Pfad zur Excel-Datei")
    ap.add_argument("--url", default=DEFAULT_URL, help="Basis-URL des Stempeluhr-Servers")
    ap.add_argument("--dry-run", action="store_true", help="Nur anzeigen, was geschrieben würde, nichts speichern")
    ap.add_argument("--no-backup", action="store_true", help="Backup-Kopie überspringen (nicht empfohlen)")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"Datei nicht gefunden: {path}", file=sys.stderr)
        return 1

    print(f"Hole Daten von {args.url}/api/{args.person}/days ...")
    try:
        days = fetch_days(args.url, args.person)
    except requests.RequestException as exc:
        print(f"Konnte Server nicht erreichen ({exc}). Ist Tailscale verbunden?", file=sys.stderr)
        return 1
    print(f"{len(days)} Tage vom Server erhalten, davon relevant (mit Inhalt):")

    relevant = {iso: d for iso, d in days.items() if has_content(d)}
    for iso in sorted(relevant):
        print(f"  {iso}")
    if not relevant:
        print("  (keine)")
        return 0

    wb = openpyxl.load_workbook(path)  # data_only=False (Standard) -> Formeln bleiben erhalten
    if SHEET not in wb.sheetnames:
        print(f"Tabellenblatt '{SHEET}' nicht gefunden. Vorhanden: {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb[SHEET]

    # Jahr aus Basisangaben!E6 lesen – ein vom Nutzer direkt eingetragener Wert
    # ("Sollarbeitszeitliste für das Jahr"), keine Formel. Bewusst NICHT über
    # Berechnung!A3 (=Formel =Basisangaben!E6): deren gecachtes Ergebnis geht
    # bei jedem openpyxl-Speichervorgang verloren (openpyxl berechnet keine
    # Formeln neu), Excel füllt es erst beim nächsten echten Öffnen wieder auf.
    if "Basisangaben" not in wb.sheetnames:
        print("Tabellenblatt 'Basisangaben' (enthält das Jahr) nicht gefunden.", file=sys.stderr)
        return 1
    sheet_year = wb["Basisangaben"]["E6"].value
    if not isinstance(sheet_year, int):
        print(f"Konnte Jahr nicht aus Basisangaben!E6 lesen (Wert: {sheet_year!r}).", file=sys.stderr)
        return 1
    print(f"\nExcel-Datei ist für Jahr {sheet_year} aufgebaut.")

    written, skipped_year, warnings_grund = [], [], []

    for iso, day in sorted(relevant.items()):
        d = dt.date.fromisoformat(iso)
        if d.year != sheet_year:
            skipped_year.append(iso)
            continue
        row = FIRST_ROW + (d - dt.date(sheet_year, 1, 1)).days

        for idx, punch in enumerate(day.get("punches", [])[:8]):
            col = COL_KOMMT_GEHT[idx]
            cell = ws.cell(row=row, column=col)
            if args.dry_run:
                print(f"  {iso} Zeile {row} {cell.coordinate} <- {punch['time']}")
            else:
                cell.value = parse_time(punch["time"])

        std = day.get("abwesenheitStd")
        if std:
            try:
                hours = float(std)
            except ValueError:
                print(f"  WARNUNG {iso}: 'abwesenheitStd' ({std!r}) ist keine Zahl, übersprungen.")
            else:
                cell = ws.cell(row=row, column=COL_ABWESENHEIT_STD)
                if args.dry_run:
                    print(f"  {iso} Zeile {row} {cell.coordinate} <- {hours} h")
                else:
                    cell.value = dt.timedelta(hours=hours)

        grund = day.get("abwesenheitGrund")
        if grund:
            if grund not in VALID_GRUENDE:
                warnings_grund.append((iso, grund))
            cell = ws.cell(row=row, column=COL_ABWESENHEIT_GRUND)
            if args.dry_run:
                print(f"  {iso} Zeile {row} {cell.coordinate} <- {grund!r}")
            else:
                cell.value = grund

        bemerkung = (day.get("bemerkung") or "").strip()
        if bemerkung:
            cell = ws.cell(row=row, column=COL_BEMERKUNG)
            if args.dry_run:
                print(f"  {iso} Zeile {row} {cell.coordinate} <- {bemerkung!r}")
            else:
                cell.value = bemerkung

        written.append(iso)

    if skipped_year:
        print(f"\nÜbersprungen (anderes Jahr als {sheet_year}): {', '.join(skipped_year)}")
    if warnings_grund:
        print("\nWARNUNG – Abwesenheitsgrund nicht in der Excel-Dropdown-Liste:")
        for iso, grund in warnings_grund:
            print(f"  {iso}: {grund!r} (erlaubt: {sorted(VALID_GRUENDE)})")

    if args.dry_run:
        print(f"\n[Dry Run] Es wurden keine Änderungen gespeichert. {len(written)} Tage wären geschrieben worden.")
        return 0

    if not args.no_backup:
        stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        # Zeitstempel VOR die Endung setzen, damit die Datei ".xlsx" bleibt und
        # Windows/Excel sie weiterhin als Excel-Datei erkennt (Doppelklick öffnet
        # sie direkt, statt als unbekannten Dateityp "BACKUP-...-Datei").
        backup_path = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
        shutil.copy2(path, backup_path)
        print(f"\nBackup gespeichert: {backup_path}")

    wb.save(path)
    print(f"{len(written)} Tage in {path.name} geschrieben.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
