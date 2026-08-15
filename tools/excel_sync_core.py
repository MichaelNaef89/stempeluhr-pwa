r"""Gemeinsame Kernlogik für den Excel-Sync – genutzt von sync_to_excel.py (CLI)
und excel_sync_gui.py (Klick-Starter-Fenster). Siehe sync_to_excel.py für die
ausführliche Beschreibung von Spalten-/Zeilen-Layout und Design-Entscheidungen.
"""

from __future__ import annotations

import datetime as dt
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import openpyxl
import requests

DEFAULT_URL = "https://pi5.tail0fe4c7.ts.net"
DEFAULT_PERSON = "michael"
DEFAULT_FILE = (
    r"C:\Users\micha\Desktop\Privat\Ensis\Zeiterfassung"
    r"\Arbeitszeiterfassung 2026 ab 01.09.2026 - Michi Näf.xlsx"
)

SHEET = "Arbeitszeiterfassung"
FIRST_ROW = 4  # entspricht dem 1. Januar des Jahres in Basisangaben!E6
COL_KOMMT_GEHT = {
    0: 7,   # G  Kommt 1
    1: 8,   # H  Geht 1
    2: 9,   # I  Kommt 2
    3: 10,  # J  Geht 2
    4: 11,  # K  Kommt 3
    5: 12,  # L  Geht 3
    6: 13,  # M  Kommt 4
    7: 14,  # N  Geht 4
}
COL_ABWESENHEIT_STD = 15    # O
COL_ABWESENHEIT_GRUND = 16  # P
COL_BEMERKUNG = 20          # T

# Muss exakt der Dropdown-Liste in P371:P379 der Excel-Vorlage entsprechen.
VALID_GRUENDE = {"Ferien", "krank", "Unfall", "geschäftlich", "Weiterbild'g"}


def fetch_days(base_url: str, person: str) -> dict:
    url = f"{base_url}/api/{person}/days"
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    return resp.json()


def parse_time(hhmm: str) -> dt.time:
    h, m = hhmm.split(":")
    return dt.time(int(h), int(m))


def has_content(day: dict) -> bool:
    return bool(day.get("punches")) or bool(day.get("abwesenheitStd")) or bool((day.get("bemerkung") or "").strip())


@dataclass
class SyncResult:
    ok: bool = False
    error: str | None = None
    relevant_isos: list[str] = field(default_factory=list)
    written: list[str] = field(default_factory=list)
    skipped_year: list[str] = field(default_factory=list)
    warnings_grund: list[tuple[str, str]] = field(default_factory=list)
    sheet_year: int | None = None
    backup_path: Path | None = None
    dry_run: bool = False


def sync_to_excel(
    person: str = DEFAULT_PERSON,
    file_path: str | Path = DEFAULT_FILE,
    base_url: str = DEFAULT_URL,
    dry_run: bool = False,
    make_backup: bool = True,
    log: Callable[[str], None] = print,
) -> SyncResult:
    """Holt die Tage eines Profils vom Server und schreibt sie in die Excel-
    Datei. `log` wird für jede Statuszeile aufgerufen (Standard: print), damit
    sowohl die CLI als auch die GUI denselben Fortschritt anzeigen können."""
    result = SyncResult(dry_run=dry_run)
    path = Path(file_path)

    if not path.exists():
        result.error = f"Datei nicht gefunden: {path}"
        log(result.error)
        return result

    log(f"Hole Daten von {base_url}/api/{person}/days ...")
    try:
        days = fetch_days(base_url, person)
    except requests.RequestException as exc:
        result.error = f"Konnte Server nicht erreichen ({exc}). Ist Tailscale verbunden?"
        log(result.error)
        return result

    relevant = {iso: d for iso, d in days.items() if has_content(d)}
    result.relevant_isos = sorted(relevant)
    log(f"{len(days)} Tage vom Server erhalten, davon relevant (mit Inhalt): {len(relevant)}")
    for iso in result.relevant_isos:
        log(f"  {iso}")
    if not relevant:
        result.ok = True
        return result

    wb = openpyxl.load_workbook(path)  # data_only=False -> Formeln bleiben erhalten
    if SHEET not in wb.sheetnames:
        result.error = f"Tabellenblatt '{SHEET}' nicht gefunden. Vorhanden: {wb.sheetnames}"
        log(result.error)
        return result
    ws = wb[SHEET]

    if "Basisangaben" not in wb.sheetnames:
        result.error = "Tabellenblatt 'Basisangaben' (enthält das Jahr) nicht gefunden."
        log(result.error)
        return result
    sheet_year = wb["Basisangaben"]["E6"].value
    if not isinstance(sheet_year, int):
        result.error = f"Konnte Jahr nicht aus Basisangaben!E6 lesen (Wert: {sheet_year!r})."
        log(result.error)
        return result
    result.sheet_year = sheet_year
    log(f"Excel-Datei ist für Jahr {sheet_year} aufgebaut.")

    for iso, day in sorted(relevant.items()):
        d = dt.date.fromisoformat(iso)
        if d.year != sheet_year:
            result.skipped_year.append(iso)
            continue
        row = FIRST_ROW + (d - dt.date(sheet_year, 1, 1)).days

        for idx, punch in enumerate(day.get("punches", [])[:8]):
            col = COL_KOMMT_GEHT[idx]
            cell = ws.cell(row=row, column=col)
            if dry_run:
                log(f"  {iso} Zeile {row} {cell.coordinate} <- {punch['time']}")
            else:
                cell.value = parse_time(punch["time"])

        std = day.get("abwesenheitStd")
        if std:
            try:
                hours = float(std)
            except ValueError:
                log(f"  WARNUNG {iso}: 'abwesenheitStd' ({std!r}) ist keine Zahl, übersprungen.")
            else:
                cell = ws.cell(row=row, column=COL_ABWESENHEIT_STD)
                if dry_run:
                    log(f"  {iso} Zeile {row} {cell.coordinate} <- {hours} h")
                else:
                    cell.value = dt.timedelta(hours=hours)

        grund = day.get("abwesenheitGrund")
        if grund:
            if grund not in VALID_GRUENDE:
                result.warnings_grund.append((iso, grund))
            cell = ws.cell(row=row, column=COL_ABWESENHEIT_GRUND)
            if dry_run:
                log(f"  {iso} Zeile {row} {cell.coordinate} <- {grund!r}")
            else:
                cell.value = grund

        bemerkung = (day.get("bemerkung") or "").strip()
        if bemerkung:
            cell = ws.cell(row=row, column=COL_BEMERKUNG)
            if dry_run:
                log(f"  {iso} Zeile {row} {cell.coordinate} <- {bemerkung!r}")
            else:
                cell.value = bemerkung

        result.written.append(iso)

    if result.skipped_year:
        log(f"Übersprungen (anderes Jahr als {sheet_year}): {', '.join(result.skipped_year)}")
    if result.warnings_grund:
        log("WARNUNG – Abwesenheitsgrund nicht in der Excel-Dropdown-Liste:")
        for iso, grund in result.warnings_grund:
            log(f"  {iso}: {grund!r} (erlaubt: {sorted(VALID_GRUENDE)})")

    if dry_run:
        log(f"[Dry Run] Es wurden keine Änderungen gespeichert. {len(result.written)} Tage wären geschrieben worden.")
        result.ok = True
        return result

    if make_backup:
        stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = path.with_name(f"{path.stem}.backup-{stamp}{path.suffix}")
        shutil.copy2(path, backup_path)
        result.backup_path = backup_path
        log(f"Backup gespeichert: {backup_path.name}")

    wb.save(path)
    log(f"{len(result.written)} Tage in {path.name} geschrieben.")
    result.ok = True
    return result
